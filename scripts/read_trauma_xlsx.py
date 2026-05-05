"""Temporal: inspeccionar Excel bolso trauma; puede borrarse tras integrar datos."""
import openpyxl
import json
import sys

path = r"C:\Users\kiwip\Downloads\Check List Bolsos de Trauma R-1(1-5).xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print("max_row", ws.max_row, "max_col", getattr(ws, "max_column", "?"))
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(tuple("" if c is None else c for c in row))
        if i >= 120:
            break
    print("---", name, f"showing first {len(rows)} rows ---")
    for r in rows:
        print(r)
