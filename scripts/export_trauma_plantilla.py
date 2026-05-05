"""Exporta plantilla bolso trauma desde Excel Forms (fila 1) a JSON."""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Misma ruta que el formulario original; override con argv[1]
DEFAULT_XLSX = Path(r"C:\Users\kiwip\Downloads\Check List Bolsos de Trauma R-1(1-5).xlsx")

RE_MIN = re.compile(r"minima:\s*(\d+)", re.I)
RE_OPT = re.compile(r"optima:\s*(\d+)", re.I)


def parse_min_opt(text: str) -> tuple[int, int]:
    t = text.replace("\xa0", " ")
    m = RE_MIN.search(t)
    o = RE_OPT.search(t)
    mn = int(m.group(1)) if m else 0
    op = int(o.group(1)) if o else mn
    return mn, op


def material_name_from_header(text: str, bolsillo_line: str | None) -> str:
    t = text.replace("\xa0", " ").strip()
    if not t:
        return ""
    lines = [ln.strip() for ln in t.split("\n") if ln.strip()]
    if not lines:
        return ""
    # Quitar líneas de cantidad del nombre
    name_lines: list[str] = []
    for ln in lines:
        if RE_MIN.search(ln) or RE_OPT.search(ln):
            break
        if ln.lower().startswith("cantidad "):
            break
        name_lines.append(ln)
    if bolsillo_line and name_lines and name_lines[0] == bolsillo_line.strip():
        name_lines = name_lines[1:]
    raw = " ".join(name_lines).strip()
    # Caso: "Mascara Bolsas para Residuos" -> separar en dos ítems no aplica en una sola columna; dejamos un nombre legible
    return raw


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.is_file():
        print(json.dumps({"error": f"No file: {path}"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Columnas de metadatos hasta el primer ítem de inventario
    start = 10
    cells = [("" if c is None else str(c)) for c in row1]

    # Detectar inicio del segundo bolso (repite Bolsillo Principal)
    end = len(cells)
    for j in range(start + 1, len(cells)):
        c = cells[j]
        if c.startswith("Bolsillo Principal") and "Kit" in c and j > start + 3:
            end = j
            break
        # Observaciones largas / duplicado
        if "Registrar cualquier anomal" in c and j > start + 20:
            end = j
            break

    ubicaciones: list[dict] = []
    current_ub: str | None = None
    mats: list[dict] = []

    def flush_ub() -> None:
        nonlocal current_ub, mats
        if current_ub and mats:
            ubicaciones.append({"nombre": current_ub, "materiales": mats})
        mats = []

    for j in range(start, end):
        raw = cells[j].strip()
        raw = re.sub(r"(?<=[^\s\n])(?=\s*Cantidad minima)", "\n", raw, flags=re.I)
        if not raw or "Registrar cualquier anomal" in raw:
            continue
        if "Ingrese aqu" in raw:
            continue

        lines = [ln.strip() for ln in raw.split("\n") if ln.strip()]
        first = lines[0] if lines else ""

        if first.startswith("Bolsillo "):
            flush_ub()
            current_ub = first
            rest = "\n".join(raw.split("\n")[1:])
            if not rest.strip():
                continue
            mn, op = parse_min_opt(raw)
            nombre = material_name_from_header(raw, first)
            if nombre:
                mats.append({"nombre": nombre, "cantidadMinima": mn, "cantidadOptima": op})
        else:
            if not current_ub:
                continue
            mn, op = parse_min_opt(raw)
            nombre = material_name_from_header(raw, None)
            if nombre:
                mats.append({"nombre": nombre, "cantidadMinima": mn, "cantidadOptima": op})

    flush_ub()

    # Un bolso = misma plantilla repetida 3 veces en la app (1..3)
    out = {
        "ubicaciones": ubicaciones,
        "bolsos": [{"numero": n, "ubicaciones": ubicaciones} for n in (1, 2, 3)],
    }
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
